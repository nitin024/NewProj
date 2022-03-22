from tkinter import *
import os.path
import datetime
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
#from PIL import Image, ImageTk


root = Tk()
root.title("Lazzy Data Extracter")

root.iconbitmap('C:\\IMEX255\\DataExtract.ico')
#root.iconbitmap('C:\\IMEX255\\pythonicon.ico')

#-- DNB Logo Code
# image_path = 'C:\\IMEX255\\DNBlogo.png'

# root.geometry('1000x700')

# image = Image.open(image_path)
# photo = ImageTk.PhotoImage(image)

# label = Label(root, image = photo)
# label.image = photo
# label.grid(column=7, row=0)
# label.pack()

result_var = ''

#Workbook instance
wb = Workbook()
#Read Spreadsheet
wb = load_workbook('LazzyDataExtracter_Output.xlsx')
ws = wb.active



#Find a way to display data from Excel
def extract_data():
    MyButton = Button(root, text="Export to Excel", command=lambda: export_to_excel())
    MyButton.grid(column=2, row=8)

    #label_result["text"] = "\n--------------------------------------------\n ** Result ** \n--------------------------------------------\n\n\n** Data Extracted 1 **\n\n** Data Extracted 2  **\n\n** Data Extracted 3 **\n\n--------------------------------------------\n"
    label_result["text"] = "\n"
    ro = 10
    for row in ws:
        co = 0
        for cell in row:
            res = Label(root,text=cell.value).grid(row=ro,column=co)
            co+=1
        ro+=1
    label_result[res]

def save_query():
    label_result["text"] = "\n--------------------------------------------\n ** Query Saved in C:\LDE**\n\n--------------------------------------------\n"

def save_template():
    label_result["text"] = "\n--------------------------------------------\n ** Saved as Template **\n\n--------------------------------------------\n"

def export_to_excel():
    label_result["text"] = "\n--------------------------------------------\n ** Exported to Excel **\n\n--------------------------------------------\n"

envLst = ["DMT1", "DEV", "SIT", "UAT", "PROD"]
applnLst = ["ASAP-LC", "DLA", "Kunde", "Sak", "Infra"]
no_of_records = ["10", "20", "30", "50", "100", "500", "All"]
GenColumnLst = ["DLA_INC-112","CBSF-123 GDPR1","Accounts", "Bruker", "Case",]
colmnCondLst = ["KUNDE_STATUS","PROSJEKTID","PROSJEKT_STATUS","KUNDENUMMER","KUNDEID","ARKIVNR","GYLDIG_TIL","KUNDETYPE"]

env_click = StringVar()
env_click.set("DEV")
sel_env_label = Label(root, text="\n Envionment: ")
sel_env_label.grid(column=0, row=0)

drop_env = OptionMenu(root, env_click, *envLst)
drop_env.grid(column=0, row=1)


appln_click = StringVar()
appln_click.set("DLA")
sel_appln_label = Label(root, text="\n Application: ")
sel_appln_label.grid(column=0, row=2)

drop = OptionMenu(root, appln_click, *applnLst)
drop.grid(column=0, row=3)

colmnCond_click1 = StringVar()
colmnCond_click1.set("KUNDE_STATUS")
colmnCond_click2 = StringVar()
colmnCond_click2.set("PROSJEKTID")
colmnCond_click3 = StringVar()
colmnCond_click3.set("PROSJEKT_STATUS")
sel_colmnCond_label = Label(root, text="\n Condition Column ")
sel_colmnCond_label.grid(column=1, row=0)
drop_colmnCond1 = OptionMenu(root, colmnCond_click1, *colmnCondLst)
drop_colmnCond1.grid(column=1, row=1, sticky="e")
drop_colmnCond2 = OptionMenu(root, colmnCond_click2, *colmnCondLst)
drop_colmnCond2.grid(column=1, row=2, sticky="e")
drop_colmnCond3 = OptionMenu(root, colmnCond_click3, *colmnCondLst)
drop_colmnCond3.grid(column=1, row=3, sticky="e")

sel_inpcolmnCond_label = Label(root, text="\n Condition Value ")
sel_inpcolmnCond_label.grid(column=2, row=0, sticky="w")
inputcolmnCond1 = Entry(root, width=10)
inputcolmnCond1.grid(column=2, row=1, sticky="w")

inputcolmnCond2 = Entry(root, width=10)
inputcolmnCond2.grid(column=2, row=2, sticky="w")

inputcolmnCond3 = Entry(root, width=10)
inputcolmnCond3.grid(column=2, row=3, sticky="w")

chk_state1= BooleanVar()
chk_state2 = BooleanVar()
chk_state3 = BooleanVar()
chk_state4 = BooleanVar()
chk_state5 = BooleanVar()

sel_appln_label = Label(root, text="\n Fields to be fetched: ")
sel_appln_label.grid(column=3, row=0, sticky="w")
chk_state1.set(True) #set check state
chk = Checkbutton(root, text='PROSJEKTID', var=chk_state1)
chk.grid(column=3, row=1, sticky="w")
chk_state2.set(False) #set check state
chk = Checkbutton(root, text='PROSJEKT_STATUS', var=chk_state2)
chk.grid(column=3, row=2, sticky="w")
chk_state3.set(False) #set check state
chk = Checkbutton(root, text='KUNDEID', var=chk_state3)
chk.grid(column=3, row=3, sticky="w")
chk_state4.set(False) #set check state
chk = Checkbutton(root, text='KUNDE_STATUS', var=chk_state4)
chk.grid(column=3, row=4, sticky="w")


appln_GenColumnLst = StringVar()
appln_GenColumnLst.set("   ")
sel_GenColumnLst_label = Label(root, text="\n Pre-Saved Template: ")
sel_GenColumnLst_label.grid(column=4, row=0)

drop_GenColumnLst = OptionMenu(root, appln_GenColumnLst, *GenColumnLst)
drop_GenColumnLst.grid(column=4, row=1)


appln_no_of_records = StringVar()
appln_no_of_records.set("10")
sel_no_of_records_label = Label(root, text="\n Number of Records: ")
sel_no_of_records_label.grid(column=5, row=0)

drop_no_of_records = OptionMenu(root, appln_no_of_records, *no_of_records)
drop_no_of_records.grid(column=5, row=1)

MyButton = Button(root, text="Save Query", command=lambda: save_query(),bg="Light Green", fg="black",font = ('Sans','10','bold'))
MyButton.grid(column=3, row=7)

MyButton = Button(root, text="Save as Template", command=lambda: save_template(),bg="Green", fg="black",font = ('Sans','10','bold'))
MyButton.grid(column=4, row=7)

MyButton = Button(root, text="Extract Data", command=lambda: extract_data(),bg="black", fg="white",font = ('Sans','10','bold'))
MyButton.grid(column=5, row=7)



Query_input_label = Label(root, text="Query Input File Name: ")
Query_input_label.grid(column=0, row=7, sticky="e")
inputQueryFile = Entry(root, width=20)
inputQueryFile.focus()
inputQueryFile.grid(column=1, row=7, sticky="w")



label_result = Label(root, text="\n** Ready to extract the Data **\n")
label_result.grid(column=1, row=9)

root.mainloop()
